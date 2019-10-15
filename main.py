import openpyxl
from openpyxl import Workbook
from tika import parser
import os
import re
                        #Cycle through each file in Tablet Scan Subfolder (PDF Versions Only)

#Global Variables / PDF Parser & Test Output
tabletScanPath = "C:\\Users\Carol\Desktop\XL Sample Sheets\\"                                       #Where to pull Tablet Scans from
dirs = os.listdir(tabletScanPath)
saveLocation = "C:\\Users\Carol\Desktop\\New Product List.xlsx"                     #Where to save the Spreadsheet
lb = "------------------------------------------------------------------------------------------------------------------"

#Global Variables / Spreadsheet
wb = Workbook()
destination_spreadsheet = 'New Product List.xlsx'
ws = wb.active
ws.title = "Product Data"
ws['A1'] = "SKU"                                                                    #Category Header, SKU
ws['B1'] = "Description"                                                            #Category Header, Description
ws['C1'] = "Product Height"                                                         #Category Header, Product Height
ws['D1'] = "Product Width"                                                          #Category Header, Product Width
ws['E1'] = "Product Depth"                                                          #Category Header, Product Depth
ws['F1'] = "Product Weight"                                                         #Category Header, Product Weight
ws['G1'] = "Boxed Height"                                                           #Category Header, Boxed Product Height
ws['H1'] = "Boxed Width"                                                            #Category Header, Boxed Product Width
ws['I1'] = "Boxed Depth"                                                            #Category Header, Boxed Product Depth
ws['J1'] = "Boxed Weight"                                                           #Category Header, Boxed Product Weight
ws['K1'] = "Internet Price"                                                         #Category Header, Product Price
ws['L1'] = "MSRP"                                                                   #Category Header, Wholesale Price
ws['M1'] = "Wholesale Price"                                                        #Category Header, MSRP

rowVal = 2

for file in dirs:
                #------- Within this block, any code is applied to each file in Test Folder --------
    pdfContent = parser.from_file(tabletScanPath + file)             #Parses PDF document
    tsContent = (pdfContent['content'])                                             #Returns only the CONTENT of the PDF

    #------------------------------------------------ REGEX Searches ---------------------------------------------------
    productSku = re.search(r'(CD..[0-9]+).*', file)                #Matches any product that matches CDXX####
    #if productSku == None:
        #productSku = re.search(r'(CD..[0-9]+).*', tsContent)         #Matches any product that matches CDXX########

    productDims = re.search(r'BOX SIZE\n\nH\s([0-9]*\.*[0-9]*)\"*i*n*\s*([0-9]*\.*[0-9]*)\"*i*n*\n\nW\s([0-9]*\.*[0-9]*)\"*i*n*\s*([0-9]*\.*[0-9]*)\"*i*n*\n\nD\s([0-9]\.*[0-9]*)\"*i*n*\s*([0-9]*)\"*i*n*', tsContent)                  #Group 1 = Product Size, Group 2 = Box Size
    try:
        productHeight = productDims.group(1)
    except AttributeError:
        '''try:
            productDims = re.search(r'BOX SIZE\n\nH\s([0-9]*).*\s*([0-9]*)\n\nW\s([0-9]*).\d*\s*([0-9]*)\n\nD\s([0-9]*).*\s([0-9]*)\n', tsContent)
            productHeight = productDims.group(1)
        except AttributeError:
            try:
                productDims = re.search(r'BOX SIZE\n\nH\s([0-9]*).*\s*([0-9]*)\n\nW\s([0-9]*).\d*\s*([0-9]*)\n\nD\s([0-9]*).*\s([0-9]*)', tsContent)
                productHeight = productDims.group(1)
            except AttributeError:'''
        productHeight = "--"

    if productHeight == None:                                                          #Error Loop, If No Box Size Found
        productHeight = "--"
        #re.search(r'H\s([0-9]+)', tsContent)

    try:
        productWidth = productDims.group(3)                   #Group 1 = Product Size, Group 2 = Box Size
    except AttributeError:
        try:
            productDims = re.search(r'BOX SIZE\n\nH\s([0-9]*\.*[0-9]*)\"*i*n*\s*([0-9]*\.*[0-9]*)\"*i*n*\n\nW\s([0-9]*\.*[0-9]*)\"*i*n*\s*([0-9]*\.*[0-9]*)\"*i*n*\n\nD\s([0-9]\.*[0-9]*)\"*i*n*\s*([0-9]*)\"*i*n*', tsContent)
            productWidth = productDims.group(3)
        except AttributeError:
            productWidth = "--"

    if productWidth == None:                                                            #Error Loop, If No Box Size Found
        productWidth = "--"
        #re.search(r'W\s([0-9]+)', tsContent)

    try:
        productDepth = productDims.group(5)                   #Group 1 = Product Size, Group 2 = Box Size
    except AttributeError:
        productDepth = "--"

    if productDepth == None:                                                            #Error Loop, If No Box Size Found
        productDepth = "--"
        #re.search(r'D\s([0-9]+)', tsContent)

    productWeight = re.search(r'LBS\s([0-9]+\.*[0-9]*\s)([0-9]+\.*[0-9]*)', tsContent)          #Group 1 = Product Weight, Group 2 = Box Weight
    if productWeight == None:
        productWeight = re.search('LBS\s([0-9]+\.*[0-9]*\s)', tsContent)                  #Error Loop 1/2, if NO DECIMAL in weight.
        if productWeight == None:
            productWeight = re.search('LBS\s([0-9]+\.*[0-9]*\s)', tsContent)                       #Error Loop 2/2, if NO Box Weight
            if productWeight == None:
                productWeight = "--"

    productDesc = re.search(r'ITEM #\n(\n[\w\W+\s]+)(CD..)', tsContent)                        #Find Product Description on Tablet Scan
    if productDesc == None:
        productDesc = re.search(r'DESCRIPTION:\s...\n\n([\w\W+\s]+)\n[\d+]+', tsContent)

    productPrice = re.search(r'TOTAL:\s([0-9]+.[0-9][0-9])\$*\s*', tsContent)           #Find Product Total Price on Tablet Scan


    #--------------------------------------------- REGEX Search Cleaning -----------------------------------------------
    try:
        cleanProductDesc = productDesc.group(1) #.split(productSku.group())            #Cut String off when SKU is found in string
        cleanProductDesc = cleanProductDesc.lstrip("ITEM #")                      #Remove 'ITEM #' from start of string
        cleanProductDesc = cleanProductDesc.strip('\n')                           #Remove any '\n'  from string
        cleanProductDesc = cleanProductDesc.replace('\n', "")                     #Remove any '\n' from center of string
    except AttributeError:
        cleanProductDesc = "No Product Description"                                 #Print if Regex search fails to find description
    except:
        cleanProductDesc = "No Product Description"

    #---------------------------------------------- XL Sheet Functions -------------------------------------------------
    #if productSku:
    ws['A' + rowVal.__str__()] = productSku.group(1)        #Store all Found SKUS in New Spreadsheet, col A
    #ws['A' + rowVal.__str__()] = file        #Store all Found SKUS in New Spreadsheet, col A

    try:
        ws['B' + rowVal.__str__()] = cleanProductDesc        #Store Descriptions.. Col B
    except AttributeError:
        ws['B' + rowVal.__str__()] = "No Product Description"           #If no description, store no description, Col B
    except IndexError:
        ws['B' + rowVal.__str__()] = "No Product Description"  # If no description, store no description, Col B
    except TypeError:
        ws['B' + rowVal.__str__()] = "No Product Description"

    if productHeight:
        ws['C' + rowVal.__str__()] = productHeight     #Store Product Height.. Col C

    if productWidth:
        try:
            ws['D' + rowVal.__str__()] = productWidth      #Store Product Width.. Col D
        except AttributeError:
            ws['D' + rowVal.__str__()] = "--"
    if productDepth:
        ws['E' + rowVal.__str__()] = productDepth      #Store Product Depth.. Col E

    if productWeight:
        try:
            ws['F' + rowVal.__str__()] = productWeight.group(1)     #Store Product Weight.. Col F
        except AttributeError:
            ws['F' + rowVal.__str__()] = "--"
    try:
        ws['G' + rowVal.__str__()] = productDims.group(2)     #Store Boxed Height.. Col G
    except IndexError:
        ws['G' + rowVal.__str__()] = "--"                       #If no Boxed Height, Insert '--'
    except AttributeError:
        ws['G' + rowVal.__str__()] = "--"

    try:
        ws['H' + rowVal.__str__()] = productDims.group(4)      #Store Boxed Width.. Col H
    except IndexError:
        ws['H' + rowVal.__str__()] = "--"                       #If no Boxed Width, Insert '--'
    except AttributeError:
        ws['H' + rowVal.__str__()] = "--"

    try:
        ws['I' + rowVal.__str__()] = productDims.group(6)      #Store Boxed Depth.. Col I
    except IndexError:
        ws['I' + rowVal.__str__()] = "--"
    except AttributeError:
        ws['I' + rowVal.__str__()] = "--"

    try:
        ws['J' + rowVal.__str__()] = productWeight.group(2)     #Store Boxed Weight.. Col J
    except IndexError:
        ws['J' + rowVal.__str__()] = "--"                       #If no Boxed Weight, Insert '--'
    except AttributeError:
        ws['J' + rowVal.__str__()] = "--"                       #If no Boxed Weight, Insert '--'

    if productPrice:
        ws['K' + rowVal.__str__()] = productPrice.group(1)      #Store Produt Price.. Col K
        msrp = float(productPrice.group(1).lstrip("$"))
        msrp = msrp * 3
        ws['L' + rowVal.__str__()] = msrp

        wholesalePrice = msrp / 2
        ws['M' + rowVal.__str__()] = wholesalePrice
    else:
        try:
            productPrice = re.search(r'\n([0-9]+.[0-9]+)\$\s*\n\nI', tsContent)
            ws['K' + rowVal.__str__()] = productPrice.group(1)      #If Product Price Regex Search Fails,
            msrp = float(productPrice.group(1).lstrip("$"))
            msrp = msrp * 3
            ws['L' + rowVal.__str__()] = msrp

            wholesalePrice = msrp / 2
            ws['M' + rowVal.__str__()] = wholesalePrice
        except AttributeError:
            try:
                productPrice = re.search(r'TOTAL: (\$[0-9]+.[0-9]+)', tsContent)
                ws['K' + rowVal.__str__()] = productPrice.group(1)
                msrp = float(productPrice.group(1).lstrip("$"))
                msrp = msrp * 3
                ws['L' + rowVal.__str__()] = msrp

                wholesalePrice = msrp / 2
                ws['M' + rowVal.__str__()] = wholesalePrice
            except AttributeError:
                try:
                    productPrice = re.search(r'TOTAL:\nW\n\n\$([0-9]*\.*[0-9]*)', tsContent)
                    ws['K' + rowVal.__str__()] = productPrice.group(1)
                    msrp = float(productPrice.group(1).lstrip("$"))
                    msrp = msrp * 3
                    ws['L' + rowVal.__str__()] = msrp
                except AttributeError:
                    ws['K' + rowVal.__str__()] = "--"



    rowVal += 1

    #--------------------------------------------------- Test Output ---------------------------------------------------
    print(tsContent)
    #print(productHeight)
    #print(productDepth)
    #print(productWidth)
    #print(productDesc)
    #print(cleanProductDesc)

    '''
    print(tsContent)
    print(lb)
    if productSku:
        #print(productSku.group(0))
        print(file)
    try:
        print(cleanProductDesc[0])
    except AttributeError:
        print(cleanProductDesc)

    print(lb)

    if productHeight:
        print(productHeight.group(1))

    if productWidth:
        print(productWidth.group(1))

    if productDepth:
        print(productDepth.group(1))

    if productWeight:
        print(productWeight.group(1))

    print(lb)

    print(lb)

    try:
        print(productHeight.group(2))
    except IndexError:
        print('--')

    try:
        print(productWidth.group(2))
    except IndexError:
        print('--')

    try:
        print(productDepth.group(2))
    except IndexError:
        print('--')

    try:
        print(productWeight.group(2))
    except IndexError:
        print('--')
    except AttributeError:
        print('--')

    print(lb)

    if productPrice:
        print(productPrice.group(1))
    else:
        productPrice = re.search(r'\n([0-9]+.[0-9]+)\$\s*\n\nI', tsContent)
        print(productPrice.group(1))

    print(lb + '\n\n\n' )
    '''
wb.save(saveLocation)                                                               #Save Workbook

print('Document saved to previously defined Desktop.')                              #Print Success Message
